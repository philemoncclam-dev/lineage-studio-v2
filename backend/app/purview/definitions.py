"""Import column definitions from a spreadsheet and push them onto Purview.

The workflow this serves: a user drills into a table, uploads the .xlsx their
data-governance team maintains by hand, and we reconcile that sheet against the
column names Purview actually holds. Spreadsheet names are written by humans and
almost never match the physical names exactly (`Customer Id` vs `customer_id`
vs `CustomerId`), so matching is fuzzy — but never silently. Every proposal
carries a confidence and a status, the user confirms, and only then do we write.

Matching uses `difflib` from the stdlib on purpose: the normalisation below
(camelCase splitting, separator folding) does the heavy lifting, and a real
fuzzy-matching dependency would earn its keep only for problems this shape is
already solved for.

Writes go through `WriteSession` like every other mutation. Preview is a dry
run, confirm is the same code path with `apply=True`, so what the user approves
is exactly what is sent.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Any, Iterable
from urllib.parse import unquote

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, Field

from .client import PurviewClient, PurviewError
from .writer import WriteResult, WriteSession

# The attribute the Purview portal surfaces as the editable description on a
# scanned asset. Column entities carry both `description` (scan-populated, read
# only in practice) and `userDescription`; writing the latter is what actually
# shows up in the UI. Verified live on a `fabric_lakehouse_table_column`.
DESCRIPTION_ATTRIBUTE = "userDescription"

# Above this a match is safe enough to pre-select for the user.
CONFIDENT = 0.88
# Below this we do not propose at all — it would be noise, not a suggestion.
PLAUSIBLE = 0.62
# Two candidates closer together than this are ambiguous: show the match but
# leave it unchecked so the user decides rather than us guessing.
AMBIGUITY_MARGIN = 0.06

_CAMEL_BOUNDARY = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")
_NON_WORD = re.compile(r"[^0-9a-z]+")


def normalise(name: str) -> str:
    """Fold a column name to the form used for comparison.

    `Customer Id`, `customer_id` and `CustomerId` all collapse to `customerid`,
    which makes the three spellings the same string rather than three fuzzy
    near-misses that would each need a threshold to survive.
    """
    spaced = _CAMEL_BOUNDARY.sub(" ", (name or "").strip())
    return _NON_WORD.sub("", spaced.lower())


def _tokens(name: str) -> list[str]:
    spaced = _CAMEL_BOUNDARY.sub(" ", (name or "").strip())
    return [t for t in _NON_WORD.split(spaced.lower()) if t]


def similarity(a: str, b: str) -> float:
    """Confidence that two column names denote the same column, 0..1.

    Identical after normalisation is 1.0. Otherwise we take the better of a
    character-level ratio and a token-order-insensitive ratio, so `order_date`
    and `Date Of Order` are recognised despite the reordering, which a plain
    character ratio scores poorly.
    """
    na, nb = normalise(a), normalise(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0

    char_ratio = SequenceMatcher(None, na, nb).ratio()
    ta, tb = sorted(_tokens(a)), sorted(_tokens(b))
    token_ratio = SequenceMatcher(None, " ".join(ta), " ".join(tb)).ratio()
    best = max(char_ratio, token_ratio)

    # A reordering of the *same* tokens is the same name, but never claim the
    # certainty of an exact match — 0.97 keeps it pre-selected yet visibly
    # distinct from a literal hit in the review list. The cap applies to any
    # non-identical pair, so 1.0 always means "these are the same name".
    if ta == tb:
        best = max(best, 0.97)
    return round(min(best, 0.97), 3)


# --- spreadsheet parsing -----------------------------------------------


@dataclass
class SheetRow:
    """One name/description pair lifted from the uploaded file."""

    name: str
    description: str


def _looks_like_header(name: str, description: str) -> bool:
    """Whether the first row names the columns rather than carrying data.

    Sheets in the wild sometimes have a header and sometimes do not, so guess
    from the wording instead of demanding one shape. Guessing wrong only costs
    one unmatched row, which stays visible to the user either way.
    """
    a, b = normalise(name), normalise(description)
    header_a = {"name", "column", "columnname", "field", "fieldname", "attribute"}
    header_b = {"description", "definition", "comment", "businessdefinition", "meaning"}
    return a in header_a or b in header_b


def _rows_from_records(records: Iterable[Iterable[Any]]) -> list[SheetRow]:
    """First cell is the name, second the description; extra columns ignored.

    Blank rows and rows with no description are dropped: an empty definition is
    nothing to import, and writing one would blank out whatever Purview holds.
    """
    rows: list[SheetRow] = []
    for record in records:
        cells = list(record)
        name = str(cells[0]).strip() if len(cells) > 0 and cells[0] is not None else ""
        desc = str(cells[1]).strip() if len(cells) > 1 and cells[1] is not None else ""
        if not name or not desc:
            continue
        if not rows and _looks_like_header(name, desc):
            continue
        rows.append(SheetRow(name=name, description=desc))
    return rows


def parse_csv(data: bytes) -> list[SheetRow]:
    text = data.decode("utf-8-sig", errors="replace")
    return _rows_from_records(csv.reader(io.StringIO(text)))


def parse_xlsx(data: bytes) -> list[SheetRow]:
    """Read the first worksheet of an .xlsx workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise DefinitionImportError(
            "openpyxl is not installed; .xlsx import is unavailable"
        ) from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return _rows_from_records(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()


class DefinitionImportError(ValueError):
    """The uploaded file could not be read as a definitions sheet."""


def parse_definitions(filename: str, data: bytes) -> list[SheetRow]:
    """Dispatch on extension; .xls (the old binary format) is not supported."""
    lower = (filename or "").lower()
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return parse_csv(data)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx(data)
    raise DefinitionImportError(
        f"Unsupported file type {filename!r} — upload .xlsx or .csv."
    )


# --- matching ----------------------------------------------------------


@dataclass
class TargetColumn:
    """A real Purview column: what a spreadsheet row can be matched onto."""

    guid: str
    name: str
    data_type: str | None = None
    current_description: str | None = None


@dataclass
class Proposal:
    """A spreadsheet row paired with the column we think it describes."""

    source_name: str
    description: str
    column_guid: str | None = None
    column_name: str | None = None
    confidence: float = 0.0
    #: exact | fuzzy | ambiguous | unmatched
    status: str = "unmatched"
    #: Pre-ticked in the review UI. Never true for anything uncertain.
    selected: bool = False
    #: Other columns that scored nearly as well, so the user can see the doubt.
    alternatives: list[str] = field(default_factory=list)


def match_rows(rows: Iterable[SheetRow], columns: Iterable[TargetColumn]) -> list[Proposal]:
    """Pair every spreadsheet row with its best column, keeping doubt visible.

    Every row comes back, including ones that matched nothing — a silently
    dropped row is indistinguishable from a row that imported fine, and the
    whole point of the confirm step is that the user can see what will happen.

    Each column can be claimed only once. Rows are resolved best-score-first so
    a strong match takes the column ahead of a weaker row that also wanted it,
    rather than the outcome depending on spreadsheet row order.
    """
    columns = list(columns)
    rows = list(rows)

    scored: list[tuple[float, int, list[tuple[float, TargetColumn]]]] = []
    for index, row in enumerate(rows):
        ranked = sorted(
            ((similarity(row.name, col.name), col) for col in columns),
            key=lambda pair: pair[0],
            reverse=True,
        )
        scored.append((ranked[0][0] if ranked else 0.0, index, ranked))

    proposals: dict[int, Proposal] = {}
    taken: set[str] = set()

    for _, index, ranked in sorted(scored, key=lambda s: -s[0]):
        row = rows[index]
        proposal = Proposal(source_name=row.name, description=row.description)

        available = [(score, col) for score, col in ranked if col.guid not in taken]
        if available and available[0][0] >= PLAUSIBLE:
            score, col = available[0]
            runner_up = available[1][0] if len(available) > 1 else 0.0
            ambiguous = score < 1.0 and (score - runner_up) < AMBIGUITY_MARGIN

            proposal.column_guid = col.guid
            proposal.column_name = col.name
            proposal.confidence = score
            # Runners-up are shown to justify doubt, so an exact hit lists none:
            # there is nothing for the user to weigh up.
            proposal.alternatives = (
                []
                if score >= 1.0
                else [c.name for s, c in available[1:4] if s >= PLAUSIBLE]
            )
            if ambiguous:
                proposal.status = "ambiguous"
            elif score >= 1.0:
                proposal.status = "exact"
            else:
                proposal.status = "fuzzy"
            proposal.selected = score >= CONFIDENT and not ambiguous
            taken.add(col.guid)

        proposals[index] = proposal

    return [proposals[i] for i in range(len(rows))]


# --- reading the target table -----------------------------------------


def _columns_with_guids(detail: dict) -> list[TargetColumn]:
    """Columns hanging off an entity's `columns` relationship, with GUIDs.

    `ingest._columns_of` drops the GUID because the graph does not need it; the
    push path does, so this keeps it. Data-type spelling varies by entity type
    (`dataType` on lakehouse table columns, `data_type` on tabular_schema ones).
    """
    entity = detail.get("entity", detail)
    referred = detail.get("referredEntities") or {}
    out: list[TargetColumn] = []
    for col in (entity.get("relationshipAttributes") or {}).get("columns") or []:
        guid = col.get("guid")
        name = col.get("displayText") or col.get("qualifiedName")
        if not guid or not name:
            continue
        attrs = (referred.get(guid) or {}).get("attributes") or {}
        out.append(
            TargetColumn(
                guid=guid,
                name=unquote(name),
                data_type=attrs.get("dataType") or attrs.get("data_type") or attrs.get("type"),
                current_description=attrs.get(DESCRIPTION_ATTRIBUTE) or attrs.get("description"),
            )
        )
    return out


def table_columns(client: PurviewClient, table_guid: str) -> list[TargetColumn]:
    """Target columns for a table, from wherever Fabric hung them.

    Same two-place lookup as the ingest path: lakehouse tables carry `columns`
    directly, warehouse views only via a `tabular_schema` entity.
    """
    detail = client.get_entity(table_guid)
    columns = _columns_with_guids(detail)
    if columns:
        return columns

    entity = detail.get("entity", detail)
    schema = (entity.get("relationshipAttributes") or {}).get("tabular_schema")
    schema_guid = (schema or {}).get("guid")
    if not schema_guid:
        return []
    return _columns_with_guids(client.get_entity(schema_guid))


# --- pushing -----------------------------------------------------------


def build_write_session(
    assignments: Iterable["Assignment"],
    client: PurviewClient | None = None,
    apply: bool = False,
) -> WriteSession:
    """Queue one partial-attribute update per accepted assignment.

    Atlas' `PUT /atlas/v2/entity/guid/{guid}?name=<attr>` updates a single
    attribute and leaves the rest of the entity alone, which matters: a full
    entity PUT would round-trip every scan-populated attribute and risk
    clobbering them.
    """
    session = WriteSession(client=client, apply=apply)
    for item in assignments:
        session.add(
            "PUT",
            f"/atlas/v2/entity/guid/{item.column_guid}?name={DESCRIPTION_ATTRIBUTE}",
            body=item.description,
            describes=f"Set description on column {item.column_name or item.column_guid}",
        )
    return session


# --- API ---------------------------------------------------------------


class ColumnOut(BaseModel):
    guid: str
    name: str
    data_type: str | None = None
    current_description: str | None = None


class ProposalOut(BaseModel):
    source_name: str
    description: str
    column_guid: str | None = None
    column_name: str | None = None
    confidence: float
    status: str
    selected: bool
    alternatives: list[str] = Field(default_factory=list)


class MatchResponse(BaseModel):
    table_guid: str
    columns: list[ColumnOut]
    proposals: list[ProposalOut]


class Assignment(BaseModel):
    column_guid: str
    column_name: str | None = None
    description: str


class ApplyRequest(BaseModel):
    assignments: list[Assignment]
    #: False previews. True still needs `PURVIEW_ALLOW_WRITE`; `WriteSession`
    #: downgrades to a dry run rather than erroring if the gate is off.
    apply: bool = False


router = APIRouter(prefix="/purview/definitions", tags=["definitions"])


@router.post("/match", response_model=MatchResponse)
async def match_definitions(
    table_guid: str = Form(...),
    file: UploadFile = File(...),
) -> MatchResponse:
    """Fuzzy-match an uploaded sheet against a table's real Purview columns."""
    data = await file.read()
    try:
        rows = parse_definitions(file.filename or "", data)
    except DefinitionImportError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="No name/description rows found — column A should hold column "
            "names and column B their descriptions.",
        )

    try:
        columns = table_columns(PurviewClient(), table_guid)
    except PurviewError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    if not columns:
        raise HTTPException(
            status_code=404, detail=f"No columns found for table {table_guid}."
        )

    return MatchResponse(
        table_guid=table_guid,
        columns=[ColumnOut(**vars(c)) for c in columns],
        proposals=[ProposalOut(**vars(p)) for p in match_rows(rows, columns)],
    )


@router.post("/apply")
def apply_definitions(req: ApplyRequest) -> dict:
    """Preview or push the confirmed descriptions onto the column entities."""
    if not req.assignments:
        raise HTTPException(status_code=400, detail="No assignments to apply.")
    try:
        session = build_write_session(req.assignments, apply=req.apply)
        result: WriteResult = session.run()
    except PurviewError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return result.to_dict()
